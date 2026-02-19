#!/usr/bin/env python3
import re
import hashlib
import csv
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
import urllib.parse
import socket
import html
import base64
import os

# Install these if you don't have them: pip install requests openpyxl
try:
    import requests
except ImportError:
    print("ERROR: 'requests' library not found!")
    print("Please install it by running: pip install requests")
    exit(1)

try:
    from openpyxl import Workbook
except ImportError:
    print("ERROR: 'openpyxl' library not found!")
    print("Please install it by running: pip install openpyxl")
    exit(1)

# AUTHORIZED USERS - Add username:password pairs here
AUTHORIZED_USERS = {
    'admin': 'password123',  # Change these!
    'user1': 'user1pass',
    'user2': 'user2pass',
}

# Or load from environment variable (for cloud deployment)
# Format: USERNAME1:PASSWORD1,USERNAME2:PASSWORD2
if os.getenv('AUTHORIZED_USERS'):
    AUTHORIZED_USERS = {}
    for user_pass in os.getenv('AUTHORIZED_USERS').split(','):
        username, password = user_pass.split(':')
        AUTHORIZED_USERS[username] = password

# CSV will be downloadable from the web interface

def clean_text(text):
    """Remove special characters and decode HTML entities"""
    if not text:
        return ""
    
    # Decode HTML entities (e.g., &amp; -> &, &quot; -> ", etc.)
    text = html.unescape(text)
    
    # Remove or replace problematic characters
    # Keep basic punctuation but remove control characters
    text = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', text)
    
    # Normalize quotes and apostrophes
    text = text.replace('"', '"').replace('"', '"')
    text = text.replace(''', "'").replace(''', "'")
    text = text.replace('–', '-').replace('—', '-')
    
    # Remove any remaining unusual Unicode characters that might cause issues
    text = text.encode('ascii', 'ignore').decode('ascii')
    
    return text.strip()

def scrape(url):
    """Scrape product information from a URL"""
    try:
        # Use requests library instead of curl
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        print(f"Scraping: {url}")
        response = requests.get(url, headers=headers, timeout=30, allow_redirects=True)
        response.raise_for_status()
        html_content = response.text
        
        # Extract title
        title = ""
        m = re.search(r'og:title["\'][^>]*content=["\']([^"\']+)["\']', html_content, re.IGNORECASE)
        if m:
            title = m.group(1)
        else:
            # Fallback to regular title tag
            m = re.search(r'<title>([^<]+)</title>', html_content, re.IGNORECASE)
            if m:
                title = m.group(1)
        
        # Extract description as HTML (preserve formatting)
        desc_html = ""
        
        # Try to find product description in common HTML structures
        # Look for description div/section with common class names
        desc_patterns = [
            r'<div[^>]*class=["\'][^"\']*description[^"\']*["\'][^>]*>(.*?)</div>',
            r'<div[^>]*class=["\'][^"\']*product-description[^"\']*["\'][^>]*>(.*?)</div>',
            r'<div[^>]*id=["\']description["\'][^>]*>(.*?)</div>',
            r'<section[^>]*class=["\'][^"\']*description[^"\']*["\'][^>]*>(.*?)</section>',
        ]
        
        for pattern in desc_patterns:
            m = re.search(pattern, html_content, re.IGNORECASE | re.DOTALL)
            if m:
                desc_html = m.group(1).strip()
                break
        
        # If no HTML description found, fallback to meta description but wrap in <p> tag
        if not desc_html:
            m = re.search(r'og:description["\'][^>]*content=["\']([^"\']+)["\']', html_content, re.IGNORECASE)
            if m:
                desc_text = html.unescape(m.group(1))
                desc_html = f"<p>{desc_text}</p>"
            else:
                # Try meta description
                m = re.search(r'<meta\s+name=["\']description["\']\s+content=["\']([^"\']+)["\']', html_content, re.IGNORECASE)
                if m:
                    desc_text = html.unescape(m.group(1))
                    desc_html = f"<p>{desc_text}</p>"
        
        # Clean up the HTML but preserve structure
        if desc_html:
            # Remove script and style tags
            desc_html = re.sub(r'<script[^>]*>.*?</script>', '', desc_html, flags=re.DOTALL | re.IGNORECASE)
            desc_html = re.sub(r'<style[^>]*>.*?</style>', '', desc_html, flags=re.DOTALL | re.IGNORECASE)
            # Remove comments
            desc_html = re.sub(r'<!--.*?-->', '', desc_html, flags=re.DOTALL)
            # Trim whitespace but keep HTML tags
            desc_html = desc_html.strip()
            # Limit length if needed
            if len(desc_html) > 2000:
                desc_html = desc_html[:2000] + "..."
        
        # Extract price
        price = ""
        # Try multiple price patterns - prioritize numeric JSON first (most accurate)
        patterns = [
            r'"price":\s*([0-9.]+)',  # JSON price as number (most reliable - matches first occurrence)
            r'"price":\s*"([^"]+)"',  # JSON price as string
            r'itemprop=["\']price["\']\s+content=["\']([^"\']+)["\']',  # Schema.org
            r'\$\s*([0-9,]+\.?[0-9]*)'  # Dollar sign (least reliable - fallback)
        ]
        for pattern in patterns:
            matches = re.findall(pattern, html_content)
            if matches:
                # For numeric JSON pattern, take first match (main product)
                # For others, also take first match
                price = matches[0]
                break
        
        # Extract compare at price (original/before sale price)
        compare_at_price = ""
        compare_patterns = [
            # BigCommerce RRP price
            r'"rrp_without_tax":\s*\{[^}]*"value":\s*([0-9.]+)',
            # Shopify compare at price - main product only (not selling plans)
            # Look for compare_at_price in main product JSON, before any selling_plan_allocations
            r'"compare_at_price":\s*"([0-9.]+)"',
            r'"compareAtPrice":\s*"([0-9.]+)"',
            # Magento old-price patterns
            r'id=["\']old-price-[^"\']*["\'][^>]*data-price-amount=["\']([0-9.]+)["\']',
            r'class=["\'][^"\']*old-price[^"\']*["\'][^>]*data-price-amount=["\']([0-9.]+)["\']',
            r'class=["\'][^"\']*regular-price[^"\']*["\'][^>]*data-price-amount=["\']([0-9.]+)["\']',
            # Magento in JSON-LD
            r'"price":\s*"([0-9.]+)"[^}]*"priceType":\s*"ListPrice"',
            # Schema.org original price
            r'property=["\']product:original_price:amount["\']\s+content=["\']([^"\']+)',
            # Common sale patterns (visible text on page)
            r'class=["\'][^"\']*rrp[^"\']*["\'][^>]*>\s*\$?([0-9.]+)',
            r'class=["\'][^"\']*was-price[^"\']*["\'][^>]*>\s*\$?([0-9.]+)',
            r'class=["\'][^"\']*original[^"\']*["\'][^>]*>\s*\$?([0-9.]+)',
            r'Was:\s*\$?([0-9.]+)',
        ]
        for pattern in compare_patterns:
            match = re.search(pattern, html_content, re.IGNORECASE)
            if match:
                val = match.group(1)
                # Only use if it's a valid price (not empty, not null, reasonable amount)
                if val and val.lower() not in ['null', '0', '0.0', '']:
                    # Check if it looks like a price (not too large, not subscription data)
                    try:
                        num = float(val)
                        if 0.01 < num < 10000:  # Reasonable price range
                            compare_at_price = val
                            break
                    except:
                        pass
        
        # Additional check: if compare_at_price equals current price, it's not a real compare at
        if compare_at_price and price:
            try:
                if float(compare_at_price) == float(price):
                    compare_at_price = ""
            except:
                pass
        
        # Extract image - prioritize product-specific images
        image = ""
        
        # Skip keywords for filtering out non-product images
        skip_keywords = ['thumbnail', '_thumb', 'placeholder', 'loading', 'icon', 'logo', 
                        '_sm.', '_xs.', '/thumb/', 'header', 'footer', 'banner', 
                        'crest', 'horizontal', 'logo', 'brand', 'site']
        
        # PRIORITY 1: og:image meta tag (most reliable for product pages)
        og_image_patterns = [
            r'<meta[^>]*property=["\']og:image["\'][^>]*content=["\'](https?://[^"\']+)["\']',
            r'<meta[^>]*content=["\'](https?://[^"\']+)["\'][^>]*property=["\']og:image["\']',
        ]
        for pattern in og_image_patterns:
            match = re.search(pattern, html_content, re.IGNORECASE)
            if match:
                img_url = match.group(1)
                # Make sure it's not a logo
                if not any(skip in img_url.lower() for skip in skip_keywords):
                    image = img_url
                    break
        
        # PRIORITY 2: JSON-LD structured data (Shopify stores use this)
        if not image:
            jsonld_patterns = [
                r'"image":\s*"(https?://[^"]+)"',
                r'"image":\s*\[\s*"(https?://[^"]+)"',
            ]
            for pattern in jsonld_patterns:
                matches = re.findall(pattern, html_content, re.IGNORECASE)
                for img_url in matches:
                    if not any(skip in img_url.lower() for skip in skip_keywords):
                        image = img_url
                        break
                if image:
                    break
        
        # PRIORITY 3: Product-specific image selectors
        if not image:
            product_patterns = [
                # Shopify product images
                r'<img[^>]*class=["\'][^"\']*product-image[^"\']*["\'][^>]*src=["\']([^"\']+)["\']',
                r'<img[^>]*class=["\'][^"\']*product__media[^"\']*["\'][^>]*src=["\']([^"\']+)["\']',
                r'<img[^>]*class=["\'][^"\']*featured-media[^"\']*["\'][^>]*src=["\']([^"\']+)["\']',
                # Main product image by ID
                r'<img[^>]*id=["\']product-image[^"\']*["\'][^>]*src=["\']([^"\']+)["\']',
                # High-res media URLs
                r'["\'](https?://[^"\']*\/media\/\d{3,4}x\d{3,4}/[^"\']+\.(?:jpg|jpeg|png|webp))["\']',
            ]
            for pattern in product_patterns:
                matches = re.findall(pattern, html_content, re.IGNORECASE)
                for img_url in matches:
                    if not any(skip in img_url.lower() for skip in skip_keywords):
                        # Skip very small images
                        if not re.search(r'_?\d{1,3}x\d{1,3}[._]', img_url):
                            image = img_url
                            break
                if image:
                    break
        
        # PRIORITY 4: srcset images (look for largest, but skip logos)
        if not image:
            srcset_pattern = r'srcset=["\']([^"\']+)["\']'
            srcset_matches = re.findall(srcset_pattern, html_content, re.IGNORECASE)
            largest_image = ""
            largest_width = 0
            
            for srcset in srcset_matches:
                sources = srcset.split(',')
                for source in sources:
                    parts = source.strip().split()
                    if len(parts) >= 2:
                        img_url = parts[0]
                        width_str = parts[1]
                        width_match = re.search(r'(\d+)w', width_str)
                        if width_match:
                            width = int(width_match.group(1))
                            # Skip logos and thumbnails, prefer largest
                            if width > largest_width and not any(skip in img_url.lower() for skip in skip_keywords):
                                largest_width = width
                                largest_image = img_url
            
            if largest_image:
                image = largest_image
        
        # PRIORITY 5: Fallback to any large image
        if not image:
            fallback_patterns = [
                r'<img[^>]*data-src=["\'](https?://[^"\']+)["\']',
                r'<img[^>]*src=["\'](https?://[^"\']+\.(?:jpg|jpeg|png|webp))["\']',
            ]
            for pattern in fallback_patterns:
                matches = re.findall(pattern, html_content, re.IGNORECASE)
                for img_url in matches:
                    if not any(skip in img_url.lower() for skip in skip_keywords):
                        # Prefer larger images
                        if re.search(r'/\d{3,4}x\d{3,4}/', img_url) or '_600x.' in img_url or '_800x.' in img_url:
                            image = img_url
                            break
                if image:
                    break
        
        # If we found a thumbnail, try to convert it to full-size
        if image and ('thumbnail' in image or re.search(r'_\d+x\d+\.', image)):
            # Try common patterns to get full-size image
            original_image = image
            
            # Pattern 1: Remove dimension suffixes: _280x280, _300x300, etc.
            full_size = re.sub(r'_\d+x\d+(\.[a-z]+)$', r'\1', image)
            if full_size != image:
                print(f"  → Upgraded image: removed dimension suffix")
                image = full_size
            
            # Pattern 2: Replace /thumbnail/ with /media/1024x1366/ (Fischer Sports pattern)
            if '/thumbnail/' in image:
                full_size = re.sub(r'/thumbnail/([^/]+/[^/]+/[^/]+)/([^_]+)_\d+x\d+(\.[a-z]+)$', 
                                  r'/media/1024x1366/\1/\2\3', image)
                if full_size != image:
                    print(f"  → Upgraded image: Fischer Sports pattern")
                    image = full_size
                else:
                    # Generic thumbnail replacement
                    full_size = re.sub(r'/thumbnail/', '/media/1024x1366/', image)
                    full_size = re.sub(r'_\d+x\d+(\.[a-z]+)$', r'\1', full_size)
                    if full_size != original_image:
                        print(f"  → Upgraded image: generic thumbnail to media")
                        image = full_size
            
            # Pattern 3: Replace dimension paths like /280x280/ with /1024x1366/
            if re.search(r'/\d+x\d+/', image):
                full_size = re.sub(r'/\d+x\d+/', '/1024x1366/', image)
                if full_size != image:
                    print(f"  → Upgraded image: replaced dimension path")
                    image = full_size
        
        # Generate SKU
        clean = re.sub(r'[^a-zA-Z]', '', title)[:5].upper().ljust(5, 'X')
        num = int(hashlib.md5(url.encode()).hexdigest(), 16) % 9000 + 1000
        sku = clean + str(num)
        
        product = {
            'sku': sku,
            'product_name': clean_text(title) or 'Unknown Product',
            'product_description': desc_html,  # Keep as HTML, don't clean
            'image_url': image,
            'variant_price': clean_text(price),
            'variant_compare_at_price': clean_text(compare_at_price) if compare_at_price else '',
            'product_url': url,
            'ratings': '',
            'scraped_at': datetime.now().strftime('%Y-%m-%d %H:%M')
        }
        
        print(f"✓ Scraped: {product['product_name'][:50]}")
        return product
        
    except requests.exceptions.RequestException as e:
        print(f"✗ Error scraping {url}: {e}")
        return {
            'sku': 'ERROR',
            'product_name': f'Error: {str(e)[:50]}',
            'product_description': '',
            'image_url': '',
            'variant_price': '',
            'variant_compare_at_price': '',
            'product_url': url,
            'ratings': '',
            'scraped_at': datetime.now().strftime('%Y-%m-%d %H:%M')
        }
    except Exception as e:
        print(f"✗ Unexpected error: {e}")
        return {
            'sku': 'ERROR',
            'product_name': 'Unexpected Error',
            'product_description': str(e),
            'image_url': '',
            'variant_price': '',
            'variant_compare_at_price': '',
            'product_url': url,
            'ratings': '',
            'scraped_at': datetime.now().strftime('%Y-%m-%d %H:%M')
        }

HTML = """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>eStreamly Product Scraper</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f7fa; min-height: 100vh; }
    .header { background: white; border-bottom: 1px solid #e5e7eb; padding: 16px 32px; display: flex; align-items: center; gap: 12px; }
        .header-logo { width: 36px; height: 36px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 8px; display: flex; align-items: center; justify-content: center; }
        .header-logo svg { width: 24px; height: 24px; }
        .header-title { font-size: 18px; font-weight: 700; color: #1a1a2e; }
        .header-subtitle { font-size: 13px; color: #6b7280; margin-left: auto; }
        .logout-btn { font-size: 13px; color: #667eea; text-decoration: none; padding: 6px 12px; border: 1px solid #e5e7eb; border-radius: 6px; transition: all 0.2s; }
        .logout-btn:hover { background: #f5f7fa; border-color: #667eea; }
        .container { max-width: 1200px; margin: 0 auto; padding: 32px; }
        .page-title { font-size: 24px; font-weight: 600; color: #111827; margin-bottom: 8px; }
        .page-description { font-size: 14px; color: #6b7280; margin-bottom: 24px; }
        .card { background: white; border: 1px solid #e5e7eb; border-radius: 8px; padding: 24px; margin-bottom: 24px; }
        .alert { background: #fef3c7; border: 1px solid #fbbf24; border-radius: 6px; padding: 12px 16px; margin-bottom: 24px; display: flex; align-items: start; gap: 12px; }
        .alert-icon { font-size: 20px; flex-shrink: 0; }
        .alert-content { flex: 1; }
        .alert-title { font-size: 14px; font-weight: 600; color: #92400e; margin-bottom: 4px; }
        .alert-text { font-size: 13px; color: #92400e; line-height: 1.5; }
        .form-label { display: block; font-size: 14px; font-weight: 500; color: #374151; margin-bottom: 8px; }
        .url-field { display: flex; gap: 8px; margin-bottom: 12px; align-items: center; }
        .url-input { flex: 1; padding: 12px 16px; border: 1px solid #d1d5db; border-radius: 6px; font-size: 14px; transition: all 0.2s; background: #f9fafb; }
        .url-input:focus { outline: none; border-color: #6366f1; background: white; box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.1); }
        .url-input::placeholder { color: #9ca3af; }
        .btn-remove { background: #fee; color: #c33; border: none; border-radius: 6px; width: 32px; height: 32px; font-size: 18px; cursor: pointer; display: flex; align-items: center; justify-content: center; transition: all 0.2s; }
        .btn-remove:hover { background: #fcc; }
        .btn-secondary { background: #f3f4f6; color: #374151; border: 1px solid #d1d5db; padding: 10px 20px; border-radius: 6px; font-size: 14px; cursor: pointer; transition: all 0.2s; margin-bottom: 16px; }
        .btn-secondary:hover { background: #e5e7eb; }
        .url-count { font-size: 13px; color: #6b7280; margin-top: 8px; }
        .btn { display: inline-flex; align-items: center; justify-content: center; gap: 8px; padding: 10px 20px; border: none; border-radius: 6px; font-size: 14px; font-weight: 500; cursor: pointer; transition: all 0.2s; width: 100%; margin-top: 16px; }
        .btn-primary { background: #6366f1; color: white; }
        .btn-primary:hover { background: #4f46e5; }
        .features { display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 16px; margin-top: 24px; }
        .feature-card { background: #f9fafb; border: 1px solid #e5e7eb; border-radius: 6px; padding: 16px; text-align: center; }
        .feature-icon { font-size: 28px; margin-bottom: 8px; }
        .feature-title { font-size: 14px; font-weight: 600; color: #111827; margin-bottom: 4px; }
        .feature-text { font-size: 13px; color: #6b7280; }
        .instructions { background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 6px; padding: 16px; margin-bottom: 24px; }
        .instructions-title { font-size: 14px; font-weight: 600; color: #0c4a6e; margin-bottom: 8px; }
        .instructions-list { font-size: 13px; color: #0c4a6e; line-height: 1.8; padding-left: 20px; }
        @media (max-width: 768px) { .container { padding: 16px; } .header { padding: 12px 16px; } .features { grid-template-columns: 1fr; } }
    </style>
</head>
<body>
    <div class="header">
        <div class="header-logo">
            <img src="https://awsmp-logos.s3.amazonaws.com/seller-a7uwjxv5o3fdo/1b5259496265e3e2da2f7f7033b49d44.png" alt="eStreamly" style="width: 24px; height: 24px; object-fit: contain; border-radius: 4px;">
        </div>
        <div class="header-title">eStreamly</div>
        <div class="header-subtitle">Product Scraper</div>
        <a href="/logout" class="logout-btn">Sign Out</a>
    </div>
    <div class="container">
        <h1 class="page-title">Product Scraper</h1>
        <p class="page-description">Extract product information from e-commerce URLs</p>
        <div class="alert">
            <div class="alert-icon">⚠️</div>
            <div class="alert-content">
                <div class="alert-title">Amazon URLs Not Supported</div>
                <div class="alert-text">This scraper does not support Amazon.com URLs. Please use URLs from other e-commerce platforms.</div>
            </div>
        </div>
        <div class="instructions">
            <div class="instructions-title">How to Use</div>
            <ol class="instructions-list">
                <li>Enter one product URL per field below</li>
                <li>Click "Add URL" for more products (up to 20)</li>
                <li>Click "Start Scraping" to extract product information</li>
                <li>Download your results as an XLSX file</li>
            </ol>
        </div>
        <div class="card">
            <form method="POST" action="/scrape" id="scrapeForm">
                <label class="form-label">Product URLs</label>
                <div id="urlFields">
                    <div class="url-field">
                        <input type="url" name="url[]" class="form-input url-input" placeholder="https://example.com/product" oninput="updateCount()">
                    </div>
                    <div class="url-field">
                        <input type="url" name="url[]" class="form-input url-input" placeholder="https://example.com/product" oninput="updateCount()">
                    </div>
                    <div class="url-field">
                        <input type="url" name="url[]" class="form-input url-input" placeholder="https://example.com/product" oninput="updateCount()">
                    </div>
                    <div class="url-field">
                        <input type="url" name="url[]" class="form-input url-input" placeholder="https://example.com/product" oninput="updateCount()">
                    </div>
                    <div class="url-field">
                        <input type="url" name="url[]" class="form-input url-input" placeholder="https://example.com/product" oninput="updateCount()">
                    </div>
                </div>
                <button type="button" class="btn btn-secondary" id="addUrlBtn" onclick="addUrlField()">+ Add URL</button>
                <div class="url-count" id="urlCount">0 URLs</div>
                <button type="submit" class="btn btn-primary">Start Scraping</button>
            </form>
        </div>
        <div class="features">
            <div class="feature-card"><div class="feature-icon">⚡</div><div class="feature-title">Fast Processing</div><div class="feature-text">Scrape multiple products quickly</div></div>
            <div class="feature-card"><div class="feature-icon">📊</div><div class="feature-title">XLSX Export</div><div class="feature-text">Download formatted spreadsheets</div></div>
            <div class="feature-card"><div class="feature-icon">🎨</div><div class="feature-title">HTML Descriptions</div><div class="feature-text">Preserves formatting in descriptions</div></div>
        </div>
    </div>
    <script>
        function updateCount() {
            const inputs = document.querySelectorAll('input[name="url[]"]');
            let count = 0;
            inputs.forEach(input => {
                if (input.value.trim().length > 0) count++;
            });
            document.getElementById('urlCount').textContent = count + ' URL' + (count !== 1 ? 's' : '');
        }
        
        function addUrlField() {
            const container = document.getElementById('urlFields');
            const currentFields = container.querySelectorAll('.url-field').length;
            if (currentFields >= 20) {
                alert('Maximum 20 URLs allowed');
                return;
            }
            const newField = document.createElement('div');
            newField.className = 'url-field';
            newField.innerHTML = '<input type="url" name="url[]" class="form-input url-input" placeholder="https://example.com/product" oninput="updateCount()"><button type="button" class="btn-remove" onclick="removeUrlField(this)" title="Remove">×</button>';
            container.appendChild(newField);
        }
        
        function removeUrlField(btn) {
            const field = btn.parentElement;
            field.remove();
            updateCount();
        }
    </script>
</body>
</html>"""

# Professional Login Page HTML
LOGIN_HTML = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sign In - eStreamly Product Scraper</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .login-container {
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            width: 100%;
            max-width: 420px;
            padding: 48px 40px;
            text-align: center;
        }
        .logo {
            width: 80px;
            height: 80px;
            margin: 0 auto 24px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            border-radius: 16px;
            display: flex;
            align-items: center;
            justify-content: center;
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.4);
        }
        .logo img {
            width: 48px;
            height: 48px;
            object-fit: contain;
            filter: brightness(0) invert(1);
        }
        .brand-name {
            font-size: 28px;
            font-weight: 700;
            color: #1a1a2e;
            margin-bottom: 8px;
            letter-spacing: -0.5px;
        }
        .product-name {
            font-size: 14px;
            color: #666;
            margin-bottom: 32px;
            font-weight: 500;
        }
        .welcome-text {
            font-size: 20px;
            font-weight: 600;
            color: #1a1a2e;
            margin-bottom: 8px;
        }
        .subtitle {
            font-size: 14px;
            color: #888;
            margin-bottom: 32px;
        }
        .form-group {
            text-align: left;
            margin-bottom: 20px;
        }
        .form-label {
            display: block;
            font-size: 13px;
            font-weight: 600;
            color: #444;
            margin-bottom: 8px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .form-input {
            width: 100%;
            padding: 14px 16px;
            border: 2px solid #e8e8e8;
            border-radius: 10px;
            font-size: 15px;
            transition: all 0.2s;
            background: #fafafa;
        }
        .form-input:focus {
            outline: none;
            border-color: #667eea;
            background: white;
            box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.1);
        }
        .form-input::placeholder {
            color: #aaa;
        }
        .signin-btn {
            width: 100%;
            padding: 16px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 10px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
            margin-top: 12px;
            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
        }
        .signin-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(102, 126, 234, 0.5);
        }
        .signin-btn:active {
            transform: translateY(0);
        }
        .error-message {
            background: #fee;
            color: #c33;
            padding: 12px 16px;
            border-radius: 8px;
            font-size: 14px;
            margin-bottom: 20px;
            border-left: 4px solid #c33;
            display: none;
        }
        .error-message.show {
            display: block;
        }
        .footer-text {
            margin-top: 32px;
            font-size: 12px;
            color: #999;
        }
        @media (max-width: 480px) {
            .login-container {
                padding: 36px 24px;
            }
            .brand-name {
                font-size: 24px;
            }
        }
    </style>
</head>
<body>
    <div class="login-container">
        <div class="logo">
            <img src="https://awsmp-logos.s3.amazonaws.com/seller-a7uwjxv5o3fdo/1b5259496265e3e2da2f7f7033b49d44.png" alt="eStreamly Logo" style="width: 48px; height: 48px; object-fit: contain;">
        </div>
        <div class="brand-name">eStreamly</div>
        <div class="product-name">Product Scraper</div>
        <div class="welcome-text">Welcome back</div>
        <div class="subtitle">Sign in to access your scraper</div>
        <div class="error-message" id="errorMsg">Invalid username or password</div>
        <form method="POST" action="/login" onsubmit="return handleSubmit(event)">
            <div class="form-group">
                <label class="form-label" for="username">Username</label>
                <input type="text" id="username" name="username" class="form-input" placeholder="Enter your username" required autofocus>
            </div>
            <div class="form-group">
                <label class="form-label" for="password">Password</label>
                <input type="password" id="password" name="password" class="form-input" placeholder="Enter your password" required>
            </div>
            <button type="submit" class="signin-btn">Sign In</button>
        </form>
        <div class="footer-text">Secure authentication</div>
    </div>
    <script>
        function handleSubmit(e) {
            return true;
        }
        // Show error if redirected with error param
        const urlParams = new URLSearchParams(window.location.search);
        if (urlParams.has('error')) {
            document.getElementById('errorMsg').classList.add('show');
        }
    </script>
</body>
</html>'''

class ScrapeHandler(BaseHTTPRequestHandler):
    # Store the last generated filename for download
    last_csv_file = None
    # Simple session storage
    active_sessions = set()
    
    def check_auth(self):
        """Check if request has valid session cookie"""
        cookie_header = self.headers.get('Cookie')
        if cookie_header:
            # Look for session cookie
            cookies = cookie_header.split(';')
            for cookie in cookies:
                cookie = cookie.strip()
                if cookie.startswith('session='):
                    session_id = cookie.split('=', 1)[1]
                    return session_id in self.active_sessions
        return False
    
    def require_auth(self):
        """Redirect to login page"""
        self.send_response(302)
        self.send_header('Location', '/login')
        self.end_headers()
    
    def set_session_cookie(self):
        """Create and set a session cookie"""
        import uuid
        session_id = str(uuid.uuid4())
        self.active_sessions.add(session_id)
        return session_id
    
    def clear_session(self):
        """Clear the session cookie"""
        cookie_header = self.headers.get('Cookie')
        if cookie_header:
            cookies = cookie_header.split(';')
            for cookie in cookies:
                cookie = cookie.strip()
                if cookie.startswith('session='):
                    session_id = cookie.split('=', 1)[1]
                    self.active_sessions.discard(session_id)
    
    def do_GET(self):
        """Serve the main page, login page, or download CSV"""
        # Public endpoints that don't require auth
        if self.path == '/login':
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(LOGIN_HTML.encode('utf-8'))
            return
        
        if self.path == '/logout':
            self.clear_session()
            self.send_response(302)
            self.send_header('Location', '/login')
            self.send_header('Set-Cookie', 'session=; Path=/; Expires=Thu, 01 Jan 1970 00:00:00 GMT')
            self.end_headers()
            return
        
        # Check authentication for protected endpoints
        if not self.check_auth():
            self.require_auth()
            return
        
        if self.path == '/':
            # Serve main page
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(HTML.encode('utf-8'))
        elif self.path.startswith('/download/'):
            # Download file (XLSX or CSV)
            filename = self.path.split('/download/')[1]
            try:
                with open(filename, 'rb') as f:
                    content = f.read()
                self.send_response(200)
                # Set correct MIME type based on file extension
                if filename.endswith('.xlsx'):
                    self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                else:
                    self.send_header('Content-Type', 'text/csv')
                self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
                self.end_headers()
                self.wfile.write(content)
            except FileNotFoundError:
                self.send_error(404, "File not found")
        else:
            self.send_error(404, "Page not found")
    
    def do_POST(self):
        """Handle login or scraping request"""
        # Handle login form submission
        if self.path == '/login':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length).decode('utf-8')
                parsed_data = urllib.parse.parse_qs(post_data)
                
                username = parsed_data.get('username', [''])[0]
                password = parsed_data.get('password', [''])[0]
                
                # Validate credentials
                if AUTHORIZED_USERS.get(username) == password:
                    # Create session
                    session_id = self.set_session_cookie()
                    self.send_response(302)
                    self.send_header('Location', '/')
                    self.send_header('Set-Cookie', f'session={session_id}; Path=/; HttpOnly')
                    self.end_headers()
                else:
                    # Invalid credentials - redirect back to login with error
                    self.send_response(302)
                    self.send_header('Location', '/login?error=1')
                    self.end_headers()
                return
            except Exception as e:
                self.send_error(400, f"Login error: {str(e)}")
                return
        
        # Check authentication for protected endpoints
        if not self.check_auth():
            self.require_auth()
            return
        
        try:
            # Read POST data
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length).decode('utf-8')
            
            # Parse URLs from multiple input fields
            parsed_data = urllib.parse.parse_qs(post_data)
            urls = []
            if 'url[]' in parsed_data:
                urls = [u.strip() for u in parsed_data['url[]'] if u.strip()]
            
            if not urls:
                self.send_error(400, "No URLs provided")
                return
            
            # Check for Amazon URLs and filter them out
            amazon_urls = []
            valid_urls = []
            for url in urls:
                if 'amazon.com' in url.lower() or 'amzn.' in url.lower():
                    amazon_urls.append(url)
                else:
                    valid_urls.append(url)
            
            if not valid_urls:
                error_html = """<!DOCTYPE html>
<html>
<head>
    <title>Error - eStreamly Scraper</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .error-box {
            background: white;
            padding: 40px;
            border-radius: 15px;
            max-width: 500px;
            text-align: center;
            box-shadow: 0 10px 30px rgba(0,0,0,0.3);
        }
        h1 {
            color: #e74c3c;
            margin-bottom: 20px;
        }
        p {
            color: #666;
            margin-bottom: 30px;
            line-height: 1.6;
        }
        a {
            display: inline-block;
            background: #667eea;
            color: white;
            padding: 12px 30px;
            border-radius: 8px;
            text-decoration: none;
            font-weight: bold;
        }
        a:hover {
            background: #5568d3;
        }
    </style>
</head>
<body>
    <div class="error-box">
        <h1>⚠️ Amazon URLs Not Supported</h1>
        <p>All the URLs you provided are from Amazon.com, which is not supported by this scraper.</p>
        <p>Please try with URLs from other e-commerce platforms like Shopify stores, WooCommerce sites, or other product pages.</p>
        <a href="/">← Go Back</a>
    </div>
</body>
</html>"""
                self.send_response(400)
                self.send_header('Content-type', 'text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write(error_html.encode('utf-8'))
                return
            
            # Scrape all URLs
            print(f"\n{'='*60}")
            if amazon_urls:
                print(f"⚠️  Skipped {len(amazon_urls)} Amazon URL(s)")
            print(f"Starting scrape of {len(valid_urls)} URL(s)")
            print(f"{'='*60}\n")
            
            products = []
            for url in valid_urls:
                product = scrape(url)
                products.append(product)
            
            # Save to XLSX
            filename = 'products_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Write headers (no formatting)
            headers = ['Sku', 'Product Name', 'Product Description', 'Image Url', 'Variant Price', 'Variant Compareat Price', 'Product Url', 'Ratings']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Write data rows
            for row_num, p in enumerate(products, 2):
                ws.cell(row=row_num, column=1, value=p['sku'])
                ws.cell(row=row_num, column=2, value=p['product_name'])
                ws.cell(row=row_num, column=3, value=p['product_description'])
                ws.cell(row=row_num, column=4, value=p['image_url'])
                ws.cell(row=row_num, column=5, value=p['variant_price'])
                ws.cell(row=row_num, column=6, value=p['variant_compare_at_price'])
                ws.cell(row=row_num, column=7, value=p['product_url'])
                ws.cell(row=row_num, column=8, value=p['ratings'])
            
            # Save workbook
            wb.save(filename)
            
            print(f"\n{'='*60}")
            print(f"✓ Saved {len(products)} products to: {filename}")
            print(f"{'='*60}\n")
            
            # Generate results page
            results_html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Scraping Results - eStreamly 🛍️ Product Scraper Pro</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}
        
        .container {{
            max-width: 900px;
            margin: 0 auto;
        }}
        
        .header {{
            text-align: center;
            color: white;
            margin-bottom: 30px;
        }}
        
        .header h1 {{
            font-size: 2.5em;
            font-weight: 700;
            margin-bottom: 10px;
            text-shadow: 0 2px 4px rgba(0,0,0,0.2);
        }}
        
        .success-banner {{
            background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
            color: white;
            padding: 25px;
            border-radius: 15px;
            margin-bottom: 25px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }}
        
        .success-banner h2 {{
            font-size: 1.5em;
            margin-bottom: 10px;
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        
        .success-banner p {{
            font-size: 1.1em;
            opacity: 0.95;
        }}
        
        .success-banner .filename {{
            background: rgba(255,255,255,0.2);
            padding: 8px 15px;
            border-radius: 8px;
            display: inline-block;
            margin-top: 10px;
            font-family: 'Courier New', monospace;
        }}
        
        .download-section {{
            background: white;
            padding: 30px;
            border-radius: 15px;
            margin-bottom: 25px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            text-align: center;
        }}
        
        .download-btn {{
            display: inline-block;
            background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
            color: white;
            padding: 18px 40px;
            border-radius: 12px;
            text-decoration: none;
            font-weight: 700;
            font-size: 1.1em;
            box-shadow: 0 4px 15px rgba(17, 153, 142, 0.4);
            transition: all 0.3s ease;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        
        .download-btn:hover {{
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(17, 153, 142, 0.6);
        }}
        
        .products-container {{
            background: white;
            padding: 30px;
            border-radius: 15px;
            margin-bottom: 25px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }}
        
        .products-header {{
            font-size: 1.3em;
            font-weight: 700;
            color: #333;
            margin-bottom: 20px;
            padding-bottom: 15px;
            border-bottom: 2px solid #f0f0f0;
        }}
        
        .product {{
            background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%);
            padding: 20px;
            margin-bottom: 15px;
            border-radius: 12px;
            border-left: 5px solid #667eea;
            transition: all 0.3s ease;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        }}
        
        .product:hover {{
            transform: translateX(5px);
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.15);
        }}
        
        .product-header {{
            display: flex;
            justify-content: space-between;
            align-items: start;
            margin-bottom: 12px;
            gap: 15px;
        }}
        
        .sku {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 6px 12px;
            border-radius: 6px;
            font-weight: 700;
            font-size: 0.85em;
            letter-spacing: 0.5px;
        }}
        
        .product-name {{
            font-size: 1.1em;
            font-weight: 600;
            color: #333;
            line-height: 1.4;
            flex: 1;
        }}
        
        .price {{
            background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
            color: white;
            padding: 8px 16px;
            border-radius: 8px;
            font-weight: 700;
            font-size: 1.1em;
            display: inline-block;
            margin-top: 8px;
        }}
        
        .url {{
            font-size: 0.85em;
            color: #999;
            margin-top: 10px;
            word-break: break-all;
            font-family: 'Courier New', monospace;
            background: #f8f9fa;
            padding: 8px 12px;
            border-radius: 6px;
        }}
        
        .back-btn {{
            display: inline-block;
            background: white;
            color: #667eea;
            padding: 15px 30px;
            border-radius: 10px;
            text-decoration: none;
            font-weight: 600;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            transition: all 0.3s ease;
        }}
        
        .back-btn:hover {{
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(0,0,0,0.15);
        }}
        
        .footer {{
            text-align: center;
            margin-top: 30px;
        }}
        
        @media (max-width: 600px) {{
            .header h1 {{
                font-size: 2em;
            }}
            
            .product-header {{
                flex-direction: column;
            }}
            
            .download-section, .products-container {{
                padding: 20px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>✓ Scraping Complete!</h1>
        </div>
        
        <div class="success-banner">
            <h2>
                <span>🎉</span>
                <span>Successfully Scraped {len(products)} Product{'s' if len(products) != 1 else ''}</span>
            </h2>
            <p>Your data has been extracted and saved</p>
            {f'<p style="margin-top: 10px; color: #fff9; font-size: 0.9em;">⚠️ Note: {len(amazon_urls)} Amazon URL(s) were skipped (not supported)</p>' if amazon_urls else ''}
            <div class="filename">📄 {filename}</div>
        </div>
        
        <div class="download-section">
            <h3 style="margin-bottom: 15px; color: #333;">Ready to Download</h3>
            <a href="/download/{filename}" class="download-btn" download>
                📥 Download XLSX File
            </a>
        </div>
        
        <div class="products-container">
            <div class="products-header">
                Product Preview ({len(products)} items)
            </div>
"""
            
            for p in products:
                results_html += f"""
            <div class="product">
                <div class="product-header">
                    <div class="product-name">{p['product_name'][:100]}</div>
                    <div class="sku">{p['sku']}</div>
                </div>
                {f'<div class="price">${p["variant_price"]}</div>' if p['variant_price'] else ''}
                <div class="url">{p['product_url'][:120]}{'...' if len(p['product_url']) > 120 else ''}</div>
            </div>
"""
            
            results_html += """
        </div>
        
        <div class="footer">
            <a href="/" class="back-btn">← Scrape More Products</a>
        </div>
    </div>
</body>
</html>"""
            
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(results_html.encode('utf-8'))
            
        except Exception as e:
            print(f"Error in POST handler: {e}")
            self.send_error(500, f"Server error: {str(e)}")
    
    def log_message(self, format, *args):
        """Suppress default server logs"""
        pass

def main():
    """Start the scraper server"""
    port = int(os.getenv('PORT', 8080))  # Use PORT env var for cloud, default 8080 for local
    
    print("=" * 60)
    print("       🛍️  eStreamly Product Scraper - Windows Ready!")
    print("=" * 60)
    print("\n📍 Server URLs:")
    print(f"   Local:   http://localhost:{port}")
    
    try:
        ip = socket.gethostbyname(socket.gethostname())
        print(f"   Network: http://{ip}:{port}")
    except:
        pass
    
    print("\n🔐 Authentication enabled - login required")
    print("\n💡 Instructions:")
    print("   1. Open the URL above in your browser")
    print("   2. Sign in with your credentials on the login page")
    print("   3. Paste product URLs (one per line)")
    print("   4. Click 'Start Scraping'")
    print("   5. Download your XLSX file")
    print("   5. Results will be saved as XLSX")
    
    print("\n⚠️  Press Ctrl+C to stop the server")
    print("=" * 60)
    print()
    
    try:
        server = HTTPServer(('0.0.0.0', port), ScrapeHandler)
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n\n" + "=" * 60)
        print("Server stopped. Goodbye!")
        print("=" * 60)

if __name__ == '__main__':
    main()
